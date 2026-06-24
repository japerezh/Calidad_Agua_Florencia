# ============================================================
# alerts.py
# Proyecto: Calidad del Agua — Caquetá
# Motor de alertas tempranas — 4 niveles por variable
# + sistema global combinado con score IA
# Ejecutar: python src/alerts.py
# ============================================================

import pandas as pd
import numpy as np
import joblib
import warnings
import os
from datetime import datetime
warnings.filterwarnings('ignore')

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.makedirs('reports/figures', exist_ok=True)
os.makedirs('reports/metrics', exist_ok=True)
os.makedirs('data/logs',       exist_ok=True)

# ============================================================
# UMBRALES AquaSense Pro — definidos por el autor
# ============================================================

UMBRALES = {
    'pH': {
        'normal':  (6.5, 8.5),
        'etiqueta_normal':  'NORMAL',
        'etiqueta_anormal': 'CRÍTICO',
        'descripcion': 'pH fuera del rango potable (6.5–8.5)'
    },
    'Oxigeno_Disuelto_mgL': {
        'normal_min': 5.0,
        'etiqueta_normal':  'NORMAL',
        'etiqueta_anormal': 'CRÍTICO',
        'descripcion': 'OD por debajo del mínimo vital (5.0 mg/L)'
    },
    'Turbidez_NTU': {
        'normal_max': 50.0,
        'etiqueta_normal':  'NORMAL',
        'etiqueta_anormal': 'CRÍTICO',
        'descripcion': 'Turbidez superior al límite (50 NTU)'
    },
    'Temperatura_C': {
        'normal':  (5.0, 35.0),
        'etiqueta_normal':  'NORMAL',
        'etiqueta_anormal': 'CRÍTICO',
        'descripcion': 'Temperatura fuera del rango admisible (5–35°C)'
    },
}

UMBRALES_INTERMEDIOS = {
    'pH': {
        'buena':  (6.8, 8.0),
        'mala':   (6.5, 6.8),
    },
    'Oxigeno_Disuelto_mgL': {
        'buena': 8.0,
        'mala':  5.0,
    },
    'Turbidez_NTU': {
        'buena': 10.0,
        'mala':  50.0,
    },
    'Temperatura_C': {
        'buena': (15.0, 25.0),
        'mala':  (5.0,  35.0),
    },
}

COLOR = {
    'NORMAL':   '#2D6A4F',
    'BUENA':    '#52B788',
    'MALA':     '#E9C46A',
    'CRÍTICO':  '#E76F51',
    'ROJO':     '#9B2226',
    'normal':   '#457B9D',
    'anomalia': '#E63946',
}

FILL_NIVEL = {
    'NORMAL':  'D9F2D0',
    'BUENA':   'EBF5EB',
    'MALA':    'FEF9E7',
    'CRÍTICO': 'FADBD8',
}

print("=" * 65)
print("  MOTOR DE ALERTAS TEMPRANAS — AquaSense Pro")
print("=" * 65)

# ============================================================
# CARGA DE MODELOS Y DATOS
# ============================================================

iso    = joblib.load('data/models/isolation_forest.pkl')
rf_clf = joblib.load('data/models/rf_clasificador.pkl')
le     = joblib.load('data/models/label_encoder.pkl')
sc_a   = joblib.load('data/models/scaler_anom.pkl')
sc_c   = joblib.load('data/models/scaler_clf.pkl')

RUTA = 'data/raw/registro_calidad_agua_PROCESADO.xlsx'
xl   = pd.ExcelFile(RUTA)
df   = pd.read_excel(RUTA, sheet_name=xl.sheet_names[0])
df['Fecha_Hora'] = pd.to_datetime(df['Fecha_Hora'])
df   = df.sort_values('Fecha_Hora').reset_index(drop=True)

# ── Feature engineering ───────────────────────────────────────
for lag in [1, 2, 4, 8, 16]:
    df[f'pH_lag{lag}']   = df['pH'].shift(lag)
    df[f'turb_lag{lag}'] = df['Turbidez_NTU'].shift(lag)
    df[f'OD_lag{lag}']   = df['Oxigeno_Disuelto_mgL'].shift(lag)
    df[f'temp_lag{lag}'] = df['Temperatura_C'].shift(lag)

df['hora_sin']   = np.sin(2*np.pi*df['Fecha_Hora'].dt.hour/24)
df['hora_cos']   = np.cos(2*np.pi*df['Fecha_Hora'].dt.hour/24)
df['dia_semana'] = df['Fecha_Hora'].dt.dayofweek

for d in [1, 4, 8]:
    df[f'turb_delta{d}'] = df['Turbidez_NTU'].diff(d)
df['pH_delta1']   = df['pH'].diff(1)
df['OD_delta1']   = df['Oxigeno_Disuelto_mgL'].diff(1)
df['temp_delta1'] = df['Temperatura_C'].diff(1)

for win in [4, 8, 16]:
    df[f'turb_mean_{win}'] = (
        df['Turbidez_NTU'].rolling(win, min_periods=win).mean())
    df[f'turb_std_{win}']  = (
        df['Turbidez_NTU'].rolling(win, min_periods=win).std())
    df[f'turb_max_{win}']  = (
        df['Turbidez_NTU'].rolling(win, min_periods=win).max())
    df[f'OD_mean_{win}']   = (
        df['Oxigeno_Disuelto_mgL'].rolling(win, min_periods=win).mean())

df['turb_log'] = np.log1p(df['Turbidez_NTU'])
df['anomalia'] = (df['Turbidez_NTU'] > 5.0).astype(int)
df_fe = df.dropna().reset_index(drop=True)

# ── Definición de feature sets (consistente con models.py) ───
EXCLUIR  = {'Fecha_Hora','Indice_Calidad_Agua','Clasificacion',
            'Temporada','anomalia','hora','dia_semana','dia_mes'}
FEATURES = [c for c in df_fe.columns if c not in EXCLUIR]

# FEATURES_ANOM: sin turbidez para evitar fuga de etiqueta
TURB_COLS     = {c for c in df_fe.columns if 'turb' in c.lower()}
EXCLUIR_ANOM  = EXCLUIR | TURB_COLS
FEATURES_ANOM = [c for c in df_fe.columns if c not in EXCLUIR_ANOM]

print(f"  Registros: {len(df_fe)} | "
      f"Período: {df_fe['Fecha_Hora'].min().date()} → "
      f"{df_fe['Fecha_Hora'].max().date()}")
print(f"  Features clf: {len(FEATURES)} | "
      f"Features anomalías: {len(FEATURES_ANOM)} (sin turbidez)\n")

# ============================================================
# SCORES IA
# ============================================================

# Clasificación: usa FEATURES completo (con turbidez)
X_all_clf = df_fe[FEATURES].values
clf_pred  = rf_clf.predict(sc_c.transform(X_all_clf))
clf_label = le.inverse_transform(clf_pred)
clf_proba = rf_clf.predict_proba(sc_c.transform(X_all_clf))

# Anomalías: usa FEATURES_ANOM (sin turbidez — sin fuga)
X_all_anom = df_fe[FEATURES_ANOM].values
sc_raw   = -iso.score_samples(sc_a.transform(X_all_anom))
sc_norm  = ((sc_raw - sc_raw.min()) /
            (sc_raw.max() - sc_raw.min() + 1e-9))

df_fe['anomaly_score']   = sc_norm
df_fe['clase_pred']      = clf_label
df_fe['prob_aceptable']  = clf_proba[:, 0]
df_fe['turb_delta1_abs'] = df_fe['turb_delta1'].abs()

# ============================================================
# EVALUACIÓN POR VARIABLE — UMBRALES AquaSense
# ============================================================

def clasificar_pH(ph):
    if ph < 6.5 or ph > 8.5:
        return 'CRÍTICO', f'pH={ph:.2f} fuera de rango (6.5–8.5)'
    elif 6.8 <= ph <= 8.0:
        return 'BUENA',   f'pH={ph:.2f} en rango óptimo'
    elif (6.5 <= ph < 6.8) or (8.0 < ph <= 8.5):
        return 'MALA',    f'pH={ph:.2f} en zona de advertencia'
    return 'NORMAL',      f'pH={ph:.2f} normal'

def clasificar_OD(do):
    if do < 5.0:
        return 'CRÍTICO', f'OD={do:.2f} mg/L bajo mínimo vital (5.0)'
    elif do < 6.0:
        return 'MALA',    f'OD={do:.2f} mg/L en zona de estrés'
    elif do < 8.0:
        return 'BUENA',   f'OD={do:.2f} mg/L aceptable'
    return 'NORMAL',      f'OD={do:.2f} mg/L óptimo'

def clasificar_turbidez(turb):
    if turb > 50.0:
        return 'CRÍTICO', f'Turbidez={turb:.2f} NTU sobre límite (50)'
    elif turb > 25.0:
        return 'MALA',    f'Turbidez={turb:.2f} NTU alta'
    elif turb > 10.0:
        return 'BUENA',   f'Turbidez={turb:.2f} NTU moderada'
    return 'NORMAL',      f'Turbidez={turb:.2f} NTU clara'

def clasificar_temperatura(temp):
    if temp < 5.0 or temp > 35.0:
        return 'CRÍTICO', f'Temp={temp:.2f}°C fuera de rango (5–35°C)'
    elif temp < 10.0 or temp > 30.0:
        return 'MALA',    f'Temp={temp:.2f}°C en estrés térmico'
    elif temp < 15.0 or temp > 25.0:
        return 'BUENA',   f'Temp={temp:.2f}°C admisible'
    return 'NORMAL',      f'Temp={temp:.2f}°C óptima'

SEVERIDAD = {'NORMAL': 0, 'BUENA': 1, 'MALA': 2, 'CRÍTICO': 3}

def nivel_global(niveles_vars):
    nivs = [SEVERIDAD[n] for n in niveles_vars]
    peor = max(nivs)
    if nivs.count(2) >= 2:
        peor = 3
    return ['NORMAL','BUENA','MALA','CRÍTICO'][peor]

def nivel_alerta_operativo(turb, delta_turb, prob_ac):
    """Capa 1 — Regla operativa basada en turbidez medida directamente."""
    delta = 0.0 if pd.isna(delta_turb) else delta_turb
    if turb > 5.0:
        return 'ROJO'
    elif delta > 1.5 or prob_ac > 0.20:
        return 'AMARILLO'
    return 'VERDE'

def nivel_alerta_ia(score, prob_ac, delta_turb):
    """Capa 2 — Experimento ML con features independientes de turbidez."""
    delta = 0.0 if pd.isna(delta_turb) else delta_turb
    if score >= 0.65 or (score >= 0.45 and prob_ac > 0.30):
        return 'ROJO'
    elif score >= 0.35 or delta > 1.5 or prob_ac > 0.20:
        return 'AMARILLO'
    return 'VERDE'

# ── Aplicar evaluación a todo el dataset ─────────────────────
resultados = []
for _, row in df_fe.iterrows():
    niv_ph,   desc_ph   = clasificar_pH(row['pH'])
    niv_do,   desc_do   = clasificar_OD(row['Oxigeno_Disuelto_mgL'])
    niv_turb, desc_turb = clasificar_turbidez(row['Turbidez_NTU'])
    niv_temp, desc_temp = clasificar_temperatura(row['Temperatura_C'])

    niv_glob = nivel_global([niv_ph, niv_do, niv_turb, niv_temp])
    niv_ia   = nivel_alerta_ia(row['anomaly_score'],
                               row['prob_aceptable'],
                               row['turb_delta1_abs'])
    niv_op   = nivel_alerta_operativo(row['Turbidez_NTU'],
                                      row['turb_delta1_abs'],
                                      row['prob_aceptable'])

    vars_alertas = []
    for niv, desc in [(niv_ph,desc_ph),(niv_do,desc_do),
                      (niv_turb,desc_turb),(niv_temp,desc_temp)]:
        if SEVERIDAD[niv] >= 2:
            vars_alertas.append(desc)

    resultados.append({
        'nivel_pH':        niv_ph,
        'nivel_OD':        niv_do,
        'nivel_turbidez':  niv_turb,
        'nivel_temp':      niv_temp,
        'nivel_global':    niv_glob,
        'nivel_IA':        niv_ia,
        'nivel_operativo': niv_op,
        'vars_en_alerta':  ' | '.join(vars_alertas) if vars_alertas
                           else 'Todas normales',
    })

df_res = pd.DataFrame(resultados)
for col in df_res.columns:
    df_fe[col] = df_res[col].values

# ============================================================
# REPORTE EN PANTALLA
# ============================================================

print("─" * 65)
print("  EVALUACIÓN POR VARIABLE — UMBRALES AquaSense Pro")
print("─" * 65)

for var, col in [('pH','nivel_pH'),('OD','nivel_OD'),
                  ('Turbidez','nivel_turbidez'),
                  ('Temperatura','nivel_temp')]:
    dist = df_fe[col].value_counts()
    print(f"\n  {var}:")
    for niv in ['NORMAL','BUENA','MALA','CRÍTICO']:
        n = dist.get(niv, 0)
        barra = '█' * int(n/len(df_fe)*50)
        print(f"    {niv:<10}: {n:>5} ({n/len(df_fe)*100:>5.1f}%)  {barra}")

print(f"\n{'─'*65}")
print("  NIVEL GLOBAL (peor variable / combinación)")
print("─" * 65)
dist_glob = df_fe['nivel_global'].value_counts()
for niv in ['NORMAL','BUENA','MALA','CRÍTICO']:
    n = dist_glob.get(niv, 0)
    barra = '█' * int(n/len(df_fe)*50)
    print(f"  {niv:<10}: {n:>5} ({n/len(df_fe)*100:>5.1f}%)  {barra}")

print(f"\n{'─'*65}")
print("  NIVEL OPERATIVO (Capa 1 — turbidez medida directamente)")
print("─" * 65)
dist_op = df_fe['nivel_operativo'].value_counts()
for niv in ['VERDE','AMARILLO','ROJO']:
    n = dist_op.get(niv, 0)
    barra = '█' * int(n/len(df_fe)*50)
    print(f"  {niv:<10}: {n:>5} ({n/len(df_fe)*100:>5.1f}%)  {barra}")

print(f"\n{'─'*65}")
print("  NIVEL IA (Capa 2 — Isolation Forest sin turbidez, experimento ML)")
print("─" * 65)
dist_ia = df_fe['nivel_IA'].value_counts()
for niv in ['VERDE','AMARILLO','ROJO']:
    n = dist_ia.get(niv, 0)
    barra = '█' * int(n/len(df_fe)*50)
    print(f"  {niv:<10}: {n:>5} ({n/len(df_fe)*100:>5.1f}%)  {barra}")

print(f"\n{'─'*65}")
print("  VERIFICACIÓN CONTRA EVENTOS REALES")
print("─" * 65)
reales    = df_fe[df_fe['anomalia']==1]
det_rojo_op  = reales[reales['nivel_operativo']=='ROJO']
det_total_op = reales[reales['nivel_operativo'].isin(['AMARILLO','ROJO'])]
det_rojo_ia  = reales[reales['nivel_IA']=='ROJO']
det_total_ia = reales[reales['nivel_IA'].isin(['AMARILLO','ROJO'])]
print(f"  Anomalías reales en dataset      : {len(reales)}")
print(f"  [Capa 1 Operativo] ROJO          : {len(det_rojo_op)} "
      f"({len(det_rojo_op)/len(reales)*100:.1f}%)")
print(f"  [Capa 1 Operativo] AMARILLO+ROJO : {len(det_total_op)} "
      f"({len(det_total_op)/len(reales)*100:.1f}%)")
print(f"  [Capa 2 ML] ROJO                 : {len(det_rojo_ia)} "
      f"({len(det_rojo_ia)/len(reales)*100:.1f}%)")
print(f"  [Capa 2 ML] AMARILLO+ROJO        : {len(det_total_ia)} "
      f"({len(det_total_ia)/len(reales)*100:.1f}%)")
print(f"  NOTA: Capa 2 con {len(FEATURES_ANOM)} features sin turbidez")

# ============================================================
# FIGURA 9: Dashboard completo con todas las variables
# ============================================================

fig = plt.figure(figsize=(16, 16))
fig.suptitle(
    'Sistema de Alertas AquaSense Pro — Calidad del Agua\n'
    'Río Hacha, Florencia, Caquetá',
    fontsize=15, fontweight='bold', y=0.98)

ax1 = fig.add_subplot(5, 1, 1)
ax1.plot(df_fe['Fecha_Hora'], df_fe['pH'],
         color='#1F4E79', linewidth=0.8)
ax1.axhspan(6.5, 8.5, alpha=0.08, color='green',
            label='Normal (6.5–8.5)')
ax1.axhline(6.5, color=COLOR['CRÍTICO'], linestyle='--',
            linewidth=0.8, alpha=0.7)
ax1.axhline(8.5, color=COLOR['CRÍTICO'], linestyle='--',
            linewidth=0.8, alpha=0.7, label='Límite AquaSense')
ax1.set_ylabel('pH')
ax1.set_title('pH — Rango normal: 6.5–8.5', fontweight='bold')
ax1.legend(loc='upper right', fontsize=8)
ax1.set_ylim(6.0, 9.0)

ax2 = fig.add_subplot(5, 1, 2)
ax2.fill_between(df_fe['Fecha_Hora'],
                 df_fe['Oxigeno_Disuelto_mgL'],
                 alpha=0.4, color='#2E86AB')
ax2.plot(df_fe['Fecha_Hora'], df_fe['Oxigeno_Disuelto_mgL'],
         color='#2E86AB', linewidth=0.8)
ax2.axhline(5.0, color=COLOR['CRÍTICO'], linestyle='--',
            linewidth=1, label='Límite crítico AquaSense (5.0)')
ax2.axhline(8.0, color=COLOR['MALA'], linestyle='--',
            linewidth=0.8, alpha=0.7, label='Óptimo (8.0)')
ax2.set_ylabel('OD (mg/L)')
ax2.set_title('Oxígeno Disuelto — Crítico: < 5.0 mg/L',
              fontweight='bold')
ax2.legend(loc='upper right', fontsize=8)

ax3 = fig.add_subplot(5, 1, 3)
mask_an = df_fe['anomalia']==1
ax3.fill_between(df_fe['Fecha_Hora'],
                 df_fe['Turbidez_NTU'],
                 alpha=0.4, color=COLOR['normal'])
ax3.plot(df_fe['Fecha_Hora'], df_fe['Turbidez_NTU'],
         color=COLOR['normal'], linewidth=0.8)
ax3.scatter(df_fe.loc[mask_an,'Fecha_Hora'],
            df_fe.loc[mask_an,'Turbidez_NTU'],
            color=COLOR['anomalia'], s=20, zorder=5,
            label='Anomalía real (turbidez > 5 NTU)')
ax3.axhline(10.0,  color=COLOR['MALA'],    linestyle='--',
            linewidth=0.8, alpha=0.8, label='Normal→Buena (10 NTU)')
ax3.axhline(50.0,  color=COLOR['CRÍTICO'], linestyle='--',
            linewidth=1, label='Límite AquaSense (50 NTU)')
ax3.set_ylabel('Turbidez (NTU)')
ax3.set_title('Turbidez — Crítico: > 50 NTU', fontweight='bold')
ax3.legend(loc='upper right', fontsize=8)

ax4 = fig.add_subplot(5, 1, 4)
ax4.fill_between(df_fe['Fecha_Hora'],
                 df_fe['Temperatura_C'],
                 alpha=0.4, color='#A23B72')
ax4.plot(df_fe['Fecha_Hora'], df_fe['Temperatura_C'],
         color='#A23B72', linewidth=0.8)
ax4.axhspan(15, 25, alpha=0.08, color='green',
            label='Óptimo (15–25°C)')
ax4.axhline(5.0,  color=COLOR['CRÍTICO'], linestyle='--',
            linewidth=0.8, alpha=0.7)
ax4.axhline(35.0, color=COLOR['CRÍTICO'], linestyle='--',
            linewidth=0.8, alpha=0.7, label='Límite AquaSense')
ax4.set_ylabel('Temperatura (°C)')
ax4.set_title('Temperatura — Crítico: < 5°C ó > 35°C',
              fontweight='bold')
ax4.legend(loc='upper right', fontsize=8)

color_ia = {'VERDE':COLOR['NORMAL'],'AMARILLO':COLOR['MALA'],
            'ROJO':COLOR['CRÍTICO']}
colores_ia = df_fe['nivel_IA'].map(color_ia)
nivel_num  = df_fe['nivel_IA'].map({'VERDE':1,'AMARILLO':2,'ROJO':3})
ax5 = fig.add_subplot(5, 1, 5)
ax5.bar(df_fe['Fecha_Hora'], nivel_num,
        color=colores_ia, width=0.01, alpha=0.85)
ax5.set_yticks([1, 2, 3])
ax5.set_yticklabels(['VERDE', 'AMARILLO', 'ROJO'], fontsize=9)
ax5.set_ylabel('Alerta IA')
ax5.set_xlabel('Fecha')
ax5.set_title(
    f'Nivel de Alerta IA — Isolation Forest '
    f'({len(FEATURES_ANOM)} features, sin turbidez)',
    fontweight='bold')
ax5.tick_params(axis='x', rotation=30)

plt.tight_layout(rect=[0, 0, 1, 0.96])
plt.savefig('reports/figures/fig9_dashboard_alertas.png', dpi=300)
plt.close()
print(f"\n  ✓ Fig 9: Dashboard completo por variable")

# ============================================================
# FIGURA 10: Tabla de referencia umbrales AquaSense
# ============================================================

fig, ax = plt.subplots(figsize=(13, 6))
ax.axis('off')

datos_tabla = [
    ['pH', '6.5 – 8.5', '6.8 – 8.0', '6.5–6.8 ó 8.0–8.5', '< 6.5 ó > 8.5'],
    ['OD (mg/L)', '> 8.0', '6.0 – 8.0', '5.0 – 6.0', '< 5.0'],
    ['Turbidez (NTU)', '≤ 10', '10 – 25', '25 – 50', '> 50'],
    ['Temperatura (°C)', '15 – 25', '10–15 ó 25–30', '5–10 ó 30–35', '< 5 ó > 35'],
]

cols_tabla = ['Parámetro', '🟢 NORMAL', '🔵 BUENA',
              '🟡 MALA', '🔴 CRÍTICO']

tabla = ax.table(
    cellText=datos_tabla,
    colLabels=cols_tabla,
    loc='center', cellLoc='center'
)
tabla.auto_set_font_size(False)
tabla.set_fontsize(11)
tabla.scale(1, 2.5)

for j in range(len(cols_tabla)):
    tabla[0, j].set_facecolor('#1F4E79')
    tabla[0, j].set_text_props(color='white', fontweight='bold')

for i in range(1, len(datos_tabla)+1):
    tabla[i, 0].set_facecolor('#F0F0F0')
    tabla[i, 0].set_text_props(fontweight='bold')
    tabla[i, 1].set_facecolor('#D9F2D0')
    tabla[i, 2].set_facecolor('#EBF5EB')
    tabla[i, 3].set_facecolor('#FEF9E7')
    tabla[i, 4].set_facecolor('#FADBD8')

ax.set_title(
    'Tabla de Referencia — Umbrales AquaSense Pro\n'
    'Clasificación de Calidad del Agua por Variable',
    fontsize=13, fontweight='bold', pad=20)

plt.tight_layout()
plt.savefig('reports/figures/fig10_umbrales_aquasense.png', dpi=300)
plt.close()
print("  ✓ Fig 10: Tabla umbrales AquaSense")

# ============================================================
# EXPORTACIÓN EXCEL
# ============================================================

wb  = Workbook()
wb.remove(wb.active)
lado  = Side(style='thin', color='BDD7EE')
borde = Border(left=lado, right=lado, top=lado, bottom=lado)

def hdr(cell, color='1F4E79'):
    cell.font      = Font(bold=True, color='FFFFFF',
                          name='Arial', size=10)
    cell.fill      = PatternFill('solid', start_color=color)
    cell.alignment = Alignment(horizontal='center',
                               vertical='center')
    cell.border    = borde

def cel(cell, bold=False, bg=None):
    cell.font      = Font(bold=bold, name='Arial', size=10)
    cell.alignment = Alignment(horizontal='center',
                               vertical='center')
    cell.border    = borde
    if bg:
        cell.fill = PatternFill('solid', start_color=bg)

def autofit(ws):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0
                 for c in col), default=0)
        ws.column_dimensions[
            get_column_letter(col[0].column)
        ].width = min(w + 4, 45)

# ── Hoja 1: Referencia umbrales AquaSense ────────────────────
ws1 = wb.create_sheet('Umbrales_AquaSense')
ws1['A1'] = 'TABLA DE REFERENCIA — UMBRALES AquaSense Pro'
ws1['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Arial')
ws1['A2'] = ('Sistema de clasificación por variable — '
             'Pérez Hurtatis, J.A. (2026)')
ws1['A2'].font = Font(size=10, italic=True, name='Arial')

h1 = ['Parámetro','Unidad','NORMAL','BUENA','MALA','CRÍTICO',
      'Acción recomendada']
for j, h in enumerate(h1, 1):
    hdr(ws1.cell(3, j, h))

datos_ref = [
    ['pH', 'Unidades',
     '6.8 – 8.0', '6.5–6.8 ó 8.0–8.5', '< 6.5 ó > 8.5', '< 6.5 ó > 8.5',
     'Verificar contaminación ácida/básica'],
    ['Oxígeno Disuelto', 'mg/L',
     '> 8.0', '6.0 – 8.0', '5.0 – 6.0', '< 5.0',
     'Revisar materia orgánica / eutrofización'],
    ['Turbidez', 'NTU',
     '≤ 10', '10 – 25', '25 – 50', '> 50',
     'Identificar fuente de sedimentos'],
    ['Temperatura', '°C',
     '15 – 25', '10–15 ó 25–30', '5–10 ó 30–35', '< 5 ó > 35',
     'Evaluar estrés térmico en ecosistema'],
]

fills_ref = ['D9F2D0', 'EBF5EB', 'FEF9E7', 'FADBD8']
for i, fila in enumerate(datos_ref, 4):
    for j, val in enumerate(fila, 1):
        bg = fills_ref[j-3] if j >= 3 and j <= 6 else 'F0F0F0'
        cel(ws1.cell(i, j, val), bg=bg, bold=(j == 1))

ws1.cell(9, 1, 'NOTA METODOLÓGICA').font = Font(
    bold=True, size=11, color='1F4E79', name='Arial')
ws1.cell(9, 1).fill = PatternFill('solid', start_color='D6E4F0')
nota = (
    f'Score de anomalía IA calculado con {len(FEATURES_ANOM)} features '
    f'independientes de turbidez (sin fuga de etiqueta). '
    f'Los umbrales AquaSense Pro clasifican cada variable físicamente. '
    f'El sistema IA complementa los umbrales absolutos detectando '
    f'anomalías relativas al comportamiento histórico del río.'
)
c = ws1.cell(10, 1, nota)
c.font = Font(name='Arial', size=10)
c.alignment = Alignment(wrap_text=True)
ws1.merge_cells('A10:G10')
ws1.row_dimensions[10].height = 60
autofit(ws1)

# ── Hoja 2: Evaluación completa por variable ─────────────────
ws2 = wb.create_sheet('Evaluación_Por_Variable')
cols_ev = ['Fecha_Hora', 'pH', 'Oxigeno_Disuelto_mgL',
           'Turbidez_NTU', 'Temperatura_C',
           'nivel_pH', 'nivel_OD', 'nivel_turbidez',
           'nivel_temp', 'nivel_global', 'nivel_IA',
           'vars_en_alerta']
hdrs_ev = ['Fecha/Hora', 'pH', 'OD (mg/L)',
           'Turbidez (NTU)', 'Temp (°C)',
           'Nivel pH', 'Nivel OD', 'Nivel Turbidez',
           'Nivel Temp', 'Nivel Global', 'Nivel IA',
           'Variables en alerta']

for j, h in enumerate(hdrs_ev, 1):
    hdr(ws2.cell(1, j, h))

for i, (_, row) in enumerate(df_fe[cols_ev].iterrows(), 2):
    niv_glob = row['nivel_global']
    niv_ia   = row['nivel_IA']
    bg_glob  = FILL_NIVEL.get(niv_glob, 'FFFFFF')

    for j, col in enumerate(cols_ev, 1):
        val = row[col]
        if isinstance(val, float): val = round(val, 3)
        c = ws2.cell(i, j,
                     str(val)[:19] if col=='Fecha_Hora' else val)
        c.border    = borde
        c.alignment = Alignment(horizontal='center',
                                vertical='center')
        c.font      = Font(name='Arial', size=9)

        if col == 'nivel_pH':
            c.fill = PatternFill('solid',
                start_color=FILL_NIVEL.get(row['nivel_pH'],'FFFFFF'))
        elif col == 'nivel_OD':
            c.fill = PatternFill('solid',
                start_color=FILL_NIVEL.get(row['nivel_OD'],'FFFFFF'))
        elif col == 'nivel_turbidez':
            c.fill = PatternFill('solid',
                start_color=FILL_NIVEL.get(row['nivel_turbidez'],'FFFFFF'))
        elif col == 'nivel_temp':
            c.fill = PatternFill('solid',
                start_color=FILL_NIVEL.get(row['nivel_temp'],'FFFFFF'))
        elif col == 'nivel_global':
            c.fill = PatternFill('solid', start_color=bg_glob)
            c.font = Font(bold=True, name='Arial', size=9)
        elif col == 'nivel_IA':
            bg_ia = {'VERDE':'D9F2D0','AMARILLO':'FEF9E7',
                     'ROJO':'FADBD8'}.get(niv_ia,'FFFFFF')
            c.fill = PatternFill('solid', start_color=bg_ia)
        else:
            c.fill = PatternFill('solid', start_color='FFFFFF')

ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols_ev))}1"
autofit(ws2)

# ── Hoja 3: Resumen distribución por variable ─────────────────
ws3 = wb.create_sheet('Distribución_Por_Variable')
ws3['A1'] = 'DISTRIBUCIÓN DE NIVELES POR VARIABLE'
ws3['A1'].font = Font(bold=True, size=13, color='1F4E79', name='Arial')

h3 = ['Variable', 'NORMAL', '% Normal', 'BUENA', '% Buena',
      'MALA', '% Mala', 'CRÍTICO', '% Crítico']
for j, h in enumerate(h3, 1):
    hdr(ws3.cell(3, j, h))

vars_info = [
    ('pH',         'nivel_pH'),
    ('OD',         'nivel_OD'),
    ('Turbidez',   'nivel_turbidez'),
    ('Temperatura','nivel_temp'),
    ('Global',     'nivel_global'),
]

for i, (nombre, col) in enumerate(vars_info, 4):
    dist = df_fe[col].value_counts()
    n    = len(df_fe)
    bg   = 'F0F0F0' if nombre == 'Global' else 'FFFFFF'
    fila = [
        nombre,
        dist.get('NORMAL',0),
        f"{dist.get('NORMAL',0)/n*100:.1f}%",
        dist.get('BUENA',0),
        f"{dist.get('BUENA',0)/n*100:.1f}%",
        dist.get('MALA',0),
        f"{dist.get('MALA',0)/n*100:.1f}%",
        dist.get('CRÍTICO',0),
        f"{dist.get('CRÍTICO',0)/n*100:.1f}%",
    ]
    fills_d = [bg,'D9F2D0','D9F2D0','EBF5EB','EBF5EB',
               'FEF9E7','FEF9E7','FADBD8','FADBD8']
    for j,(val,bg_d) in enumerate(zip(fila,fills_d),1):
        cel(ws3.cell(i,j,val), bold=(nombre=='Global'), bg=bg_d)
autofit(ws3)

# ── Hoja 4: Eventos con alertas activas ──────────────────────
ws4 = wb.create_sheet('Eventos_Con_Alerta')
mask_alerta = df_fe['nivel_global'].isin(['MALA','CRÍTICO']) | \
              df_fe['nivel_IA'].isin(['AMARILLO','ROJO'])

df_alertas = df_fe[mask_alerta][[
    'Fecha_Hora','pH','Oxigeno_Disuelto_mgL',
    'Turbidez_NTU','Temperatura_C',
    'Indice_Calidad_Agua','nivel_global',
    'nivel_IA','vars_en_alerta'
]].copy()

ws4['A1'] = f'EVENTOS CON ALERTA ACTIVA — Total: {len(df_alertas)}'
ws4['A1'].font = Font(bold=True, size=12, color='1F4E79', name='Arial')

h4 = ['Fecha/Hora','pH','OD','Turbidez','Temp',
      'ICA','Nivel Global','Nivel IA','Variables en alerta']
for j, h in enumerate(h4, 1):
    hdr(ws4.cell(2, j, h), '9B2226')

for i, (_, row) in enumerate(df_alertas.iterrows(), 3):
    niv_ia  = row['nivel_IA']
    niv_gl  = row['nivel_global']
    bg = ('FADBD8' if niv_ia=='ROJO' or niv_gl=='CRÍTICO'
          else 'FEF9E7')
    vals = [str(row['Fecha_Hora'])[:19],
            round(row['pH'],2),
            round(row['Oxigeno_Disuelto_mgL'],2),
            round(row['Turbidez_NTU'],2),
            round(row['Temperatura_C'],2),
            round(row['Indice_Calidad_Agua'],1),
            niv_gl, niv_ia, row['vars_en_alerta']]
    for j, val in enumerate(vals, 1):
        cel(ws4.cell(i, j, val), bg=bg)

autofit(ws4)

RUTA_XL = 'reports/metrics/reporte_alertas.xlsx'
wb.save(RUTA_XL)

# ── Log de eventos ────────────────────────────────────────────
ts       = datetime.now().strftime('%Y%m%d_%H%M%S')
ruta_log = f'data/logs/alertas_{ts}.txt'
with open(ruta_log, 'w', encoding='utf-8') as f:
    f.write("=" * 65 + "\n")
    f.write("  LOG DE ALERTAS — AquaSense Pro\n")
    f.write(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    f.write(f"  Registros evaluados: {len(df_fe)}\n")
    f.write(f"  Features IA (sin turbidez): {len(FEATURES_ANOM)}\n")
    f.write("=" * 65 + "\n\n")
    f.write("UMBRALES AquaSense Pro:\n")
    f.write("  pH      : NORMAL 6.5–8.5 | CRÍTICO <6.5 o >8.5\n")
    f.write("  OD      : NORMAL >5.0 mg/L | CRÍTICO <5.0 mg/L\n")
    f.write("  Turbidez: NORMAL <50 NTU | CRÍTICO >50 NTU\n")
    f.write("  Temp    : NORMAL 5–35°C | CRÍTICO <5 o >35°C\n\n")
    f.write("DISTRIBUCIÓN NIVEL GLOBAL:\n")
    for niv in ['NORMAL','BUENA','MALA','CRÍTICO']:
        n = dist_glob.get(niv, 0)
        f.write(f"  {niv:<10}: {n} ({n/len(df_fe)*100:.1f}%)\n")
    f.write("\nDISTRIBUCIÓN NIVEL IA (sin turbidez):\n")
    for niv in ['VERDE','AMARILLO','ROJO']:
        n = dist_ia.get(niv, 0)
        f.write(f"  {niv:<10}: {n} ({n/len(df_fe)*100:.1f}%)\n")
    f.write(f"\nRECALL ANOMALÍAS REALES [Capa 1 Operativo]: "
            f"{len(det_total_op)/len(reales)*100:.1f}%\n")
    f.write(f"RECALL ANOMALÍAS REALES [Capa 2 ML]: "
            f"{len(det_total_ia)/len(reales)*100:.1f}%\n")

print(f"\n{'='*65}")
print("  ARCHIVOS EXPORTADOS")
print(f"{'='*65}")
print(f"\n  FIGURAS:")
print(f"  · reports/figures/fig9_dashboard_alertas.png")
print(f"  · reports/figures/fig10_umbrales_aquasense.png")
print(f"\n  EXCEL: {RUTA_XL}")
print(f"    - Hoja 1: Umbrales AquaSense (tabla referencia)")
print(f"    - Hoja 2: Evaluación completa por variable")
print(f"    - Hoja 3: Distribución por variable")
print(f"    - Hoja 4: Eventos con alerta activa")
print(f"\n  LOG: {ruta_log}")
print(f"\n  ✓ ALERTS.PY COMPLETO")
print(f"  Ejecuta ahora: python src/validation.py")