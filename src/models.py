# ============================================================
# models.py
# Proyecto: Calidad del Agua — Caquetá
# Contiene: Bloques 5, 6, 7 + exportación completa
# Ejecutar: python src/models.py
# Outputs:
#   reports/metrics/resultados_modelos.xlsx
#   reports/figures/*.png (8 gráficas)
# ============================================================

import pandas as pd
import numpy as np
import joblib
import warnings
import os
warnings.filterwarnings('ignore')

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns

from sklearn.ensemble import (RandomForestClassifier,
                               GradientBoostingClassifier,
                               RandomForestRegressor,
                               GradientBoostingRegressor,
                               IsolationForest)
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.model_selection import (train_test_split,
                                     StratifiedKFold,
                                     cross_validate)
from sklearn.svm import OneClassSVM
from sklearn.metrics import (classification_report,
                              confusion_matrix,
                              roc_auc_score, f1_score,
                              mean_squared_error,
                              mean_absolute_error, r2_score,
                              average_precision_score,
                              precision_recall_curve,
                              make_scorer)
from xgboost import XGBClassifier, XGBRegressor
from imblearn.over_sampling import SMOTE
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side)
from openpyxl.utils import get_column_letter

# ── Crear carpetas de salida ──────────────────────────────────
os.makedirs('reports/figures', exist_ok=True)
os.makedirs('reports/metrics', exist_ok=True)
os.makedirs('data/models',     exist_ok=True)

# ── Estilo global de gráficas ─────────────────────────────────
plt.rcParams.update({
    'figure.dpi':        150,
    'font.family':       'sans-serif',
    'font.size':         11,
    'axes.titlesize':    13,
    'axes.titleweight':  'bold',
    'axes.spines.top':   False,
    'axes.spines.right': False,
    'savefig.bbox':      'tight',
    'savefig.facecolor': 'white',
})

# ── Paleta de colores del proyecto ───────────────────────────
COLOR = {
    'RF':   '#1F4E79',
    'XGB':  '#2E86AB',
    'GB':   '#A23B72',
    'VERDE':    '#2D6A4F',
    'AMARILLO': '#E9C46A',
    'ROJO':     '#E76F51',
    'normal':   '#457B9D',
    'anomalia': '#E63946',
}

# ============================================================
# RECONSTRUCCIÓN DEL DATASET
# ============================================================

RUTA = 'data/raw/registro_calidad_agua_PROCESADO.xlsx'

xl   = pd.ExcelFile(RUTA)
hoja = xl.sheet_names[0]
df   = pd.read_excel(RUTA, sheet_name=hoja)
df['Fecha_Hora'] = pd.to_datetime(df['Fecha_Hora'])
df   = df.sort_values('Fecha_Hora').reset_index(drop=True)

df['hora_sin']   = np.sin(2 * np.pi * df['Fecha_Hora'].dt.hour / 24)
df['hora_cos']   = np.cos(2 * np.pi * df['Fecha_Hora'].dt.hour / 24)
df['dia_semana'] = df['Fecha_Hora'].dt.dayofweek
df['dia_mes']    = df['Fecha_Hora'].dt.day

for lag in [1, 2, 4, 8, 16]:
    df[f'pH_lag{lag}']   = df['pH'].shift(lag)
    df[f'turb_lag{lag}'] = df['Turbidez_NTU'].shift(lag)
    df[f'OD_lag{lag}']   = df['Oxigeno_Disuelto_mgL'].shift(lag)
    df[f'temp_lag{lag}'] = df['Temperatura_C'].shift(lag)

for d in [1, 4, 8]:
    df[f'turb_delta{d}'] = df['Turbidez_NTU'].diff(d)
df['pH_delta1']   = df['pH'].diff(1)
df['OD_delta1']   = df['Oxigeno_Disuelto_mgL'].diff(1)
df['temp_delta1'] = df['Temperatura_C'].diff(1)

for win in [4, 8, 16]:
    df[f'turb_mean_{win}'] = (
        df['Turbidez_NTU'].rolling(win, min_periods=win).mean())
    df[f'turb_std_{win}']  = (
        df['Turbidez_NTU'].rolling(win, min_periods=win).std())
    df[f'turb_max_{win}']  = (
        df['Turbidez_NTU'].rolling(win, min_periods=win).max())
    df[f'OD_mean_{win}']   = (
        df['Oxigeno_Disuelto_mgL'].rolling(win,
                                            min_periods=win).mean())

df['turb_log'] = np.log1p(df['Turbidez_NTU'])
df['anomalia'] = (df['Turbidez_NTU'] > 5.0).astype(int)
df_fe = df.dropna().reset_index(drop=True)

# ── Features para clasificación y regresión (con turbidez) ───
EXCLUIR  = {
    'Fecha_Hora', 'Indice_Calidad_Agua', 'Clasificacion',
    'Temporada', 'anomalia', 'hora', 'dia_semana', 'dia_mes'
}
FEATURES = [c for c in df_fe.columns if c not in EXCLUIR]

# ── Features para detección de anomalías (SIN turbidez) ──────
# Excluye Turbidez_NTU y todas sus derivadas para evitar fuga
# de etiqueta: anomalia = (Turbidez_NTU > 5.0)
TURB_COLS     = {c for c in df_fe.columns if 'turb' in c.lower()}
EXCLUIR_ANOM  = EXCLUIR | TURB_COLS
FEATURES_ANOM = [c for c in df_fe.columns if c not in EXCLUIR_ANOM]

le           = LabelEncoder()
X            = df_fe[FEATURES].values
y_clf        = le.fit_transform(df_fe['Clasificacion'].values)
split_idx    = int(len(df_fe) * 0.80)
y_anom       = df_fe['anomalia'].values
contaminacion= y_anom.sum() / len(y_anom)

X_train_clf, X_test_clf, y_train_clf, y_test_clf = train_test_split(
    X, y_clf, test_size=0.20, random_state=42, stratify=y_clf)
scaler_clf    = StandardScaler()
X_train_clf_s = scaler_clf.fit_transform(X_train_clf)
X_test_clf_s  = scaler_clf.transform(X_test_clf)

X_train_reg   = X[:split_idx];  X_test_reg  = X[split_idx:]
y_train_reg   = df_fe['Indice_Calidad_Agua'].values[:split_idx]
y_test_reg    = df_fe['Indice_Calidad_Agua'].values[split_idx:]
scaler_reg    = StandardScaler()
X_train_reg_s = scaler_reg.fit_transform(X_train_reg)
X_test_reg_s  = scaler_reg.transform(X_test_reg)

# ── Split anomalías usando FEATURES_ANOM (sin turbidez) ──────
X_anom        = df_fe[FEATURES_ANOM].values
X_train_anom  = X_anom[:split_idx];  X_test_anom = X_anom[split_idx:]
y_train_anom  = y_anom[:split_idx]
y_test_anom   = y_anom[split_idx:]
scaler_anom   = StandardScaler()
X_train_anom_s= scaler_anom.fit_transform(X_train_anom)
X_test_anom_s = scaler_anom.transform(X_test_anom)

print(f"  Dataset: {len(df_fe)} registros | "
      f"{len(FEATURES)} features clf/reg | "
      f"{len(FEATURES_ANOM)} features anomalías | "
      f"{y_anom.sum()} anomalías ({contaminacion*100:.1f}%)\n")

# ============================================================
# HELPERS DE EXPORTACIÓN
# ============================================================

def estilo_header(cell, color='1F4E79'):
    cell.font      = Font(bold=True, color='FFFFFF',
                          name='Arial', size=10)
    cell.fill      = PatternFill('solid', start_color=color)
    cell.alignment = Alignment(horizontal='center',
                               vertical='center')
    cell.border    = borde_fino()

def estilo_celda(cell, bold=False, center=True):
    cell.font      = Font(bold=bold, name='Arial', size=10)
    cell.alignment = Alignment(
        horizontal='center' if center else 'left',
        vertical='center')
    cell.border    = borde_fino()

def borde_fino():
    lado = Side(style='thin', color='BDD7EE')
    return Border(left=lado, right=lado,
                  top=lado,  bottom=lado)

def autofit(ws):
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value else 0
             for c in col), default=0)
        ws.column_dimensions[
            get_column_letter(col[0].column)
        ].width = min(max_len + 4, 40)

# ============================================================
# BLOQUE 5 — CLASIFICACIÓN
# ============================================================

print("=" * 65)
print("  BLOQUE 5 — CLASIFICACIÓN")
print("=" * 65)

modelos_clf = {
    'Random Forest': RandomForestClassifier(
        n_estimators=200, class_weight='balanced',
        min_samples_leaf=3, random_state=42, n_jobs=-1),
    'XGBoost': XGBClassifier(
        n_estimators=200, max_depth=5, learning_rate=0.05,
        subsample=0.8, colsample_bytree=0.8,
        use_label_encoder=False, eval_metric='mlogloss',
        random_state=42, n_jobs=-1, verbosity=0),
    'Gradient Boosting': GradientBoostingClassifier(
        n_estimators=150, learning_rate=0.05,
        max_depth=4, subsample=0.8, random_state=42)
}

cv      = StratifiedKFold(n_splits=5, shuffle=True,
                           random_state=42)
scoring = {
    'f1_macro':    make_scorer(f1_score, average='macro'),
    'f1_weighted': make_scorer(f1_score, average='weighted'),
    'accuracy':    'accuracy'
}
res_clf = {}
cms_clf = {}
reports_clf = {}

for nombre, modelo in modelos_clf.items():
    modelo.fit(X_train_clf_s, y_train_clf)
    y_pred  = modelo.predict(X_test_clf_s)
    y_proba = modelo.predict_proba(X_test_clf_s)
    f1_mac  = f1_score(y_test_clf, y_pred, average='macro')
    f1_wei  = f1_score(y_test_clf, y_pred, average='weighted')
    acc     = (y_pred == y_test_clf).mean()
    roc     = roc_auc_score(y_test_clf, y_proba,
                            multi_class='ovr', average='macro')
    cv_res  = cross_validate(modelo, X_train_clf_s,
                              y_train_clf, cv=cv,
                              scoring=scoring,
                              return_train_score=True)
    gap     = (cv_res['train_f1_macro'].mean() -
               cv_res['test_f1_macro'].mean())
    cm      = confusion_matrix(y_test_clf, y_pred)
    rep     = classification_report(
        y_test_clf, y_pred,
        target_names=le.classes_,
        output_dict=True, digits=4)

    res_clf[nombre] = {
        'modelo': modelo, 'f1_macro': f1_mac,
        'f1_weighted': f1_wei, 'accuracy': acc,
        'roc_auc': roc,
        'cv_mean': cv_res['test_f1_macro'].mean(),
        'cv_std':  cv_res['test_f1_macro'].std(),
        'gap': gap, 'preds': y_pred, 'proba': y_proba
    }
    cms_clf[nombre]     = cm
    reports_clf[nombre] = rep
    print(f"  {nombre}: F1-macro={f1_mac:.4f} "
          f"ROC-AUC={roc:.4f} Gap={gap:.4f}")

mejor_clf = max(res_clf, key=lambda k: res_clf[k]['f1_macro'])
print(f"  Mejor modelo: {mejor_clf}\n")

# ── FIGURA 1: Tabla comparativa clasificación ─────────────────
fig, ax = plt.subplots(figsize=(10, 3))
ax.axis('off')
nombres = list(res_clf.keys())
cols    = ['Modelo','Accuracy','F1-macro','F1-weighted',
           'ROC-AUC','CV F1-macro','Gap']
filas   = []
for n in nombres:
    r = res_clf[n]
    marca = ' ★' if n == mejor_clf else ''
    filas.append([
        n + marca,
        f"{r['accuracy']:.4f}",
        f"{r['f1_macro']:.4f}",
        f"{r['f1_weighted']:.4f}",
        f"{r['roc_auc']:.4f}",
        f"{r['cv_mean']:.4f} ± {r['cv_std']:.4f}",
        f"{r['gap']:.4f}"
    ])
tabla = ax.table(cellText=filas, colLabels=cols,
                 loc='center', cellLoc='center')
tabla.auto_set_font_size(False)
tabla.set_fontsize(10)
tabla.scale(1, 2)
for j in range(len(cols)):
    tabla[0, j].set_facecolor('#1F4E79')
    tabla[0, j].set_text_props(color='white', fontweight='bold')
for i, n in enumerate(nombres, 1):
    color = '#D6E4F0' if n == mejor_clf else '#F8F9FA'
    for j in range(len(cols)):
        tabla[i, j].set_facecolor(color)
ax.set_title('Comparación de Modelos — Clasificación del Estado del Agua',
             fontsize=13, fontweight='bold', pad=15)
plt.tight_layout()
plt.savefig('reports/figures/fig1_comparacion_clasificacion.png', dpi=300)
plt.close()
print("  ✓ Fig 1: Tabla comparativa clasificación")

# ── FIGURA 2: Matrices de confusión ──────────────────────────
fig, axes = plt.subplots(1, 3, figsize=(15, 4))
colores_clf = [COLOR['RF'], COLOR['XGB'], COLOR['GB']]
for ax, nombre, color in zip(axes, nombres, colores_clf):
    cm  = cms_clf[nombre]
    cm_pct = cm.astype(float) / cm.sum(axis=1, keepdims=True) * 100
    sns.heatmap(cm, annot=True, fmt='d', ax=ax,
                cmap=sns.light_palette(color, as_cmap=True),
                xticklabels=le.classes_,
                yticklabels=le.classes_,
                linewidths=0.5, cbar=False)
    ax.set_title(nombre, fontweight='bold')
    ax.set_xlabel('Predicho')
    ax.set_ylabel('Real')
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            pct = cm_pct[i, j]
            ax.text(j + 0.5, i + 0.72,
                    f'({pct:.1f}%)',
                    ha='center', va='center',
                    fontsize=8, color='gray')
fig.suptitle('Matrices de Confusión — Clasificación',
             fontsize=14, fontweight='bold', y=1.02)
plt.tight_layout()
plt.savefig('reports/figures/fig2_matrices_confusion.png')
plt.close()
print("  ✓ Fig 2: Matrices de confusión")

# ── FIGURA 3: Importancia de features (Random Forest) ────────
rf_model = res_clf['Random Forest']['modelo']
feat_imp  = pd.Series(
    rf_model.feature_importances_, index=FEATURES
).sort_values(ascending=True).tail(15)

fig, ax = plt.subplots(figsize=(9, 6))
bars = ax.barh(feat_imp.index, feat_imp.values,
               color=COLOR['RF'], alpha=0.85,
               edgecolor='white', linewidth=0.5)
for bar, val in zip(bars, feat_imp.values):
    ax.text(val + 0.001, bar.get_y() + bar.get_height()/2,
            f'{val:.4f}', va='center', fontsize=9)
ax.set_xlabel('Importancia relativa')
ax.set_title('Top 15 Features más Importantes\n'
             '(Random Forest — Clasificación)',
             fontweight='bold')
ax.set_xlim(0, feat_imp.values.max() * 1.2)
plt.tight_layout()
plt.savefig('reports/figures/fig3_importancia_features.png')
plt.close()
print("  ✓ Fig 3: Importancia de features")

joblib.dump(rf_model,   'data/models/rf_clasificador.pkl')
joblib.dump(scaler_clf, 'data/models/scaler_clf.pkl')
joblib.dump(le,         'data/models/label_encoder.pkl')
print("  ✓ Bloque 5 completado\n")

# ============================================================
# BLOQUE 6 — REGRESIÓN ICA
# ============================================================

print("=" * 65)
print("  BLOQUE 6 — REGRESIÓN ICA")
print("=" * 65)

modelos_reg = {
    'Random Forest': RandomForestRegressor(
        n_estimators=200, min_samples_leaf=3,
        random_state=42, n_jobs=-1),
    'XGBoost': XGBRegressor(
        n_estimators=200, learning_rate=0.05,
        max_depth=5, subsample=0.8, colsample_bytree=0.8,
        random_state=42, n_jobs=-1, verbosity=0),
    'Gradient Boosting': GradientBoostingRegressor(
        n_estimators=150, learning_rate=0.05,
        max_depth=4, subsample=0.8, random_state=42)
}

res_reg  = {}
preds_reg= {}

for nombre, modelo in modelos_reg.items():
    modelo.fit(X_train_reg_s, y_train_reg)
    y_pred_r = modelo.predict(X_test_reg_s)
    rmse = np.sqrt(mean_squared_error(y_test_reg, y_pred_r))
    mae  = mean_absolute_error(y_test_reg, y_pred_r)
    r2   = r2_score(y_test_reg, y_pred_r)
    mape = np.mean(np.abs(
        (y_test_reg - y_pred_r) / (y_test_reg + 1e-9)
    )) * 100
    mask_crit = y_test_reg < 85
    rmse_crit = (np.sqrt(mean_squared_error(
        y_test_reg[mask_crit], y_pred_r[mask_crit]))
        if mask_crit.sum() > 0 else np.nan)

    res_reg[nombre]  = {
        'modelo': modelo, 'rmse': rmse, 'mae': mae,
        'r2': r2, 'mape': mape, 'rmse_crit': rmse_crit
    }
    preds_reg[nombre] = y_pred_r
    print(f"  {nombre}: RMSE={rmse:.4f} R²={r2:.4f}")

mejor_reg = max(res_reg, key=lambda k: res_reg[k]['r2'])
print(f"  Mejor modelo: {mejor_reg}\n")

# ── FIGURA 4: Predicción vs Real ──────────────────────────────
fig, axes = plt.subplots(1, 3, figsize=(15, 5))
fechas_test = df_fe['Fecha_Hora'].iloc[split_idx:].values
colores_reg = [COLOR['RF'], COLOR['XGB'], COLOR['GB']]

for ax, nombre, color in zip(axes, list(res_reg.keys()),
                              colores_reg):
    y_pred_r = preds_reg[nombre]
    ax.plot(fechas_test, y_test_reg,
            color='#2C3E50', linewidth=1.2,
            label='Real', alpha=0.8)
    ax.plot(fechas_test, y_pred_r,
            color=color, linewidth=1.2,
            linestyle='--', label='Predicho', alpha=0.8)
    r2 = res_reg[nombre]['r2']
    ax.set_title(f"{nombre}\nR²={r2:.4f}",
                 fontweight='bold')
    ax.set_xlabel('Fecha')
    ax.set_ylabel('ICA')
    ax.legend(fontsize=9)
    ax.tick_params(axis='x', rotation=30)

fig.suptitle('Predicción vs Real — Índice de Calidad del Agua',
             fontsize=13, fontweight='bold')
plt.tight_layout()
plt.savefig('reports/figures/fig4_prediccion_vs_real.png')
plt.close()
print("  ✓ Fig 4: Predicción vs Real")

# ── FIGURA 5: Scatter predicho vs real ───────────────────────
fig, axes = plt.subplots(1, 3, figsize=(14, 5))
for ax, nombre, color in zip(axes, list(res_reg.keys()),
                              colores_reg):
    y_pred_r = preds_reg[nombre]
    ax.scatter(y_test_reg, y_pred_r,
               alpha=0.4, color=color, s=15)
    lims = [min(y_test_reg.min(), y_pred_r.min()) - 1,
            max(y_test_reg.max(), y_pred_r.max()) + 1]
    ax.plot(lims, lims, 'k--', linewidth=1, alpha=0.6)
    ax.set_xlabel('ICA Real')
    ax.set_ylabel('ICA Predicho')
    r2   = res_reg[nombre]['r2']
    rmse = res_reg[nombre]['rmse']
    ax.set_title(f"{nombre}\nR²={r2:.4f}  RMSE={rmse:.4f}",
                 fontweight='bold')

fig.suptitle('Dispersión Predicho vs Real — ICA',
             fontsize=13, fontweight='bold')
plt.tight_layout()
plt.savefig('reports/figures/fig5_scatter_prediccion.png')
plt.close()
print("  ✓ Fig 5: Scatter predicción vs real")

xgb_reg = res_reg['XGBoost']['modelo']
joblib.dump(xgb_reg,    'data/models/xgb_regresor.pkl')
joblib.dump(scaler_reg, 'data/models/scaler_reg.pkl')
print("  ✓ Bloque 6 completado\n")

# ============================================================
# BLOQUE 7 — DETECCIÓN DE ANOMALÍAS (SIN FUGA DE TURBIDEZ)
# ============================================================

print("=" * 65)
print("  BLOQUE 7 — DETECCIÓN DE ANOMALÍAS")
print(f"  Features: {len(FEATURES_ANOM)} (turbidez excluida para evitar fuga)")
print("=" * 65)

res_anom  = {}
scores_dict = {}

# ── Isolation Forest ─────────────────────────────────────────
iso = IsolationForest(
    n_estimators=200, contamination=contaminacion,
    max_samples='auto', random_state=42, n_jobs=-1)
iso.fit(X_train_anom_s)
y_pred_iso = (iso.predict(X_test_anom_s) == -1).astype(int)
y_scr_iso  = -iso.score_samples(X_test_anom_s)
y_scr_iso  = ((y_scr_iso - y_scr_iso.min()) /
              (y_scr_iso.max() - y_scr_iso.min() + 1e-9))
ap_iso  = average_precision_score(y_test_anom, y_scr_iso)
roc_iso = roc_auc_score(y_test_anom, y_scr_iso)
f1_iso  = f1_score(y_test_anom, y_pred_iso, zero_division=0)
cm_iso  = confusion_matrix(y_test_anom, y_pred_iso)
res_anom['Isolation Forest'] = {
    'ap': ap_iso, 'roc': roc_iso, 'f1': f1_iso,
    'cm': cm_iso, 'tipo': 'No supervisado',
    'preds': y_pred_iso, 'scores': y_scr_iso
}
scores_dict['Isolation Forest'] = y_scr_iso
print(f"  Isolation Forest: AP={ap_iso:.4f} "
      f"ROC={roc_iso:.4f} F1={f1_iso:.4f}")

# ── RF + SMOTE ───────────────────────────────────────────────
k_smote       = min(3, max(1, y_train_anom.sum() - 1))
smote         = SMOTE(random_state=42, k_neighbors=k_smote)
X_tr_sm, y_tr_sm = smote.fit_resample(X_train_anom_s,
                                       y_train_anom)
rf_anom = RandomForestClassifier(
    n_estimators=200, class_weight='balanced',
    random_state=42, n_jobs=-1)
rf_anom.fit(X_tr_sm, y_tr_sm)
y_scr_rf_a   = rf_anom.predict_proba(X_test_anom_s)[:, 1]
prec_a, rec_a, thr_a = precision_recall_curve(
    y_test_anom, y_scr_rf_a)
f1_a      = 2*prec_a*rec_a/(prec_a+rec_a+1e-9)
best_thr  = thr_a[np.argmax(f1_a[:-1])]
y_pred_opt= (y_scr_rf_a >= best_thr).astype(int)
ap_rf  = average_precision_score(y_test_anom, y_scr_rf_a)
roc_rf = roc_auc_score(y_test_anom, y_scr_rf_a)
f1_rf  = f1_score(y_test_anom, y_pred_opt, zero_division=0)
cm_rf  = confusion_matrix(y_test_anom, y_pred_opt)
res_anom['RF + SMOTE'] = {
    'ap': ap_rf, 'roc': roc_rf, 'f1': f1_rf,
    'cm': cm_rf, 'tipo': 'Supervisado',
    'preds': y_pred_opt, 'scores': y_scr_rf_a,
    'umbral': best_thr
}
scores_dict['RF + SMOTE'] = y_scr_rf_a
print(f"  RF + SMOTE:       AP={ap_rf:.4f} "
      f"ROC={roc_rf:.4f} F1={f1_rf:.4f}")

# ── One-Class SVM ────────────────────────────────────────────
X_normal = X_train_anom_s[y_train_anom == 0]
oc_svm   = OneClassSVM(
    kernel='rbf', nu=contaminacion, gamma='scale')
oc_svm.fit(X_normal)
y_pred_svm = (oc_svm.predict(X_test_anom_s) == -1).astype(int)
y_scr_svm  = -oc_svm.decision_function(X_test_anom_s)
y_scr_svm  = ((y_scr_svm - y_scr_svm.min()) /
              (y_scr_svm.max() - y_scr_svm.min() + 1e-9))
ap_svm  = average_precision_score(y_test_anom, y_scr_svm)
roc_svm = roc_auc_score(y_test_anom, y_scr_svm)
f1_svm  = f1_score(y_test_anom, y_pred_svm, zero_division=0)
cm_svm  = confusion_matrix(y_test_anom, y_pred_svm)
res_anom['One-Class SVM'] = {
    'ap': ap_svm, 'roc': roc_svm, 'f1': f1_svm,
    'cm': cm_svm, 'tipo': 'Semi-supervisado',
    'preds': y_pred_svm, 'scores': y_scr_svm
}
scores_dict['One-Class SVM'] = y_scr_svm
print(f"  One-Class SVM:    AP={ap_svm:.4f} "
      f"ROC={roc_svm:.4f} F1={f1_svm:.4f}")

mejor_anom = max(res_anom, key=lambda k: res_anom[k]['f1'])
print(f"  Mejor modelo: {mejor_anom}\n")

# ── FIGURA 6: Scores de anomalía en el tiempo ────────────────
scores_all = -iso.score_samples(
    scaler_anom.transform(df_fe[FEATURES_ANOM].values))
scores_all = ((scores_all - scores_all.min()) /
              (scores_all.max() - scores_all.min() + 1e-9))

fig, axes = plt.subplots(3, 1, figsize=(14, 10), sharex=True)

ax1 = axes[0]
ax1.fill_between(df_fe['Fecha_Hora'], df_fe['Turbidez_NTU'],
                 alpha=0.4, color=COLOR['normal'])
ax1.plot(df_fe['Fecha_Hora'], df_fe['Turbidez_NTU'],
         color=COLOR['normal'], linewidth=0.8)
anomalias_mask = df_fe['anomalia'] == 1
ax1.scatter(df_fe.loc[anomalias_mask, 'Fecha_Hora'],
            df_fe.loc[anomalias_mask, 'Turbidez_NTU'],
            color=COLOR['anomalia'], s=20, zorder=5,
            label='Anomalía real')
ax1.axhline(5.0, color=COLOR['ROJO'],
            linestyle='--', linewidth=1,
            label='Umbral (5 NTU)')
ax1.set_ylabel('Turbidez (NTU)')
ax1.set_title('Serie Temporal — Turbidez y Detección de Anomalías',
              fontweight='bold')
ax1.legend(loc='upper right', fontsize=9)

ax2 = axes[1]
ax2.fill_between(df_fe['Fecha_Hora'], scores_all,
                 alpha=0.3, color=COLOR['XGB'])
ax2.plot(df_fe['Fecha_Hora'], scores_all,
         color=COLOR['XGB'], linewidth=0.8)
ax2.axhline(0.65, color=COLOR['ROJO'],
            linestyle='--', linewidth=1, label='Umbral ROJO')
ax2.axhline(0.35, color=COLOR['AMARILLO'],
            linestyle='--', linewidth=1, label='Umbral AMARILLO')
ax2.set_ylabel('Score Anomalía (sin turbidez)')
ax2.set_title('Score de Anomalía — Isolation Forest (features independientes)',
              fontweight='bold')
ax2.set_ylim(0, 1.05)
ax2.legend(loc='upper right', fontsize=9)

ax3 = axes[2]
colores_ica = df_fe['Clasificacion'].map({
    'EXCELENTE': COLOR['VERDE'],
    'BUENA':     COLOR['AMARILLO'],
    'ACEPTABLE': COLOR['ROJO']
})
ax3.scatter(df_fe['Fecha_Hora'],
            df_fe['Indice_Calidad_Agua'],
            c=colores_ica, s=4, alpha=0.6)
ax3.set_ylabel('ICA')
ax3.set_xlabel('Fecha')
ax3.set_title('Índice de Calidad del Agua en el Tiempo',
              fontweight='bold')
parches = [
    mpatches.Patch(color=COLOR['VERDE'],    label='EXCELENTE'),
    mpatches.Patch(color=COLOR['AMARILLO'], label='BUENA'),
    mpatches.Patch(color=COLOR['ROJO'],     label='ACEPTABLE'),
]
ax3.legend(handles=parches, loc='lower right', fontsize=9)
ax3.tick_params(axis='x', rotation=30)

plt.tight_layout()
plt.savefig('reports/figures/fig6_anomalias_tiempo.png')
plt.close()
print("  ✓ Fig 6: Anomalías en el tiempo")

# ── FIGURA 7: Curvas Precision-Recall ────────────────────────
fig, ax = plt.subplots(figsize=(7, 6))
colores_anom = {
    'Isolation Forest': COLOR['RF'],
    'RF + SMOTE':       COLOR['XGB'],
    'One-Class SVM':    COLOR['GB']
}
for nombre, color in colores_anom.items():
    scr = scores_dict[nombre]
    p, r, _ = precision_recall_curve(y_test_anom, scr)
    ap  = res_anom[nombre]['ap']
    ax.plot(r, p, color=color, linewidth=2,
            label=f"{nombre} (AP={ap:.3f})")
ax.set_xlabel('Recall')
ax.set_ylabel('Precision')
ax.set_title('Curvas Precision-Recall\nDetección de Anomalías (sin turbidez)',
             fontweight='bold')
ax.legend(fontsize=10)
ax.set_xlim(0, 1.02)
ax.set_ylim(0, 1.05)
ax.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig('reports/figures/fig7_precision_recall.png')
plt.close()
print("  ✓ Fig 7: Curvas Precision-Recall")

# ── FIGURA 8: Matrices de confusión anomalías ─────────────────
fig, axes = plt.subplots(1, 3, figsize=(13, 4))
nombres_anom = list(res_anom.keys())
for ax, nombre, color in zip(axes, nombres_anom,
                              list(colores_anom.values())):
    cm = res_anom[nombre]['cm']
    sns.heatmap(cm, annot=True, fmt='d', ax=ax,
                cmap=sns.light_palette(color, as_cmap=True),
                xticklabels=['Normal', 'Anomalía'],
                yticklabels=['Normal', 'Anomalía'],
                linewidths=0.5, cbar=False)
    f1  = res_anom[nombre]['f1']
    tip = res_anom[nombre]['tipo']
    ax.set_title(f"{nombre}\n{tip} — F1={f1:.4f}",
                 fontweight='bold')
    ax.set_xlabel('Predicho')
    ax.set_ylabel('Real')
fig.suptitle('Matrices de Confusión — Detección de Anomalías (sin turbidez)',
             fontsize=13, fontweight='bold')
plt.tight_layout()
plt.savefig('reports/figures/fig8_confusion_anomalias.png')
plt.close()
print("  ✓ Fig 8: Matrices de confusión anomalías")

joblib.dump(iso,        'data/models/isolation_forest.pkl')
joblib.dump(rf_anom,    'data/models/rf_anomalias.pkl')
joblib.dump(scaler_anom,'data/models/scaler_anom.pkl')
print("  ✓ Bloque 7 completado\n")

# ============================================================
# EXPORTACIÓN EXCEL — REPORTE COMPLETO
# ============================================================

print("=" * 65)
print("  EXPORTANDO EXCEL — reports/metrics/resultados_modelos.xlsx")
print("=" * 65)

wb = Workbook()
wb.remove(wb.active)

AZUL_OSC = '1F4E79'
AZUL_MED = '2E86AB'
AZUL_CLA = 'D6E4F0'
VERDE    = 'D9F2D0'
ROJO_CLA = 'FADBD8'

# ── HOJA 1: Resumen ejecutivo ─────────────────────────────────
ws1 = wb.create_sheet('Resumen')
ws1['A1'] = 'REPORTE DE RESULTADOS — PIPELINE IA CALIDAD DEL AGUA'
ws1['A1'].font = Font(bold=True, size=14,
                      color=AZUL_OSC, name='Arial')
ws1['A2'] = f'Dataset: {len(df_fe)} registros | ' \
            f'{len(FEATURES)} features clf/reg | ' \
            f'{len(FEATURES_ANOM)} features anomalías | ' \
            f'Período: {df_fe["Fecha_Hora"].min().date()} → ' \
            f'{df_fe["Fecha_Hora"].max().date()}'
ws1['A2'].font = Font(size=11, name='Arial', italic=True)

ws1['A3'] = ('NOTA: Features de anomalías excluyen turbidez y derivadas '
             'para evitar fuga de etiqueta (anomalia = Turbidez > 5.0 NTU)')
ws1['A3'].font = Font(size=10, name='Arial', color='C0392B', italic=True)

fila = 5
titulos_seccion = [
    ('CLASIFICACIÓN DEL ESTADO DEL AGUA', AZUL_OSC),
    ('REGRESIÓN DEL ICA', AZUL_MED),
    ('DETECCIÓN DE ANOMALÍAS', '6B4C8A'),
]
for titulo, color in titulos_seccion:
    ws1.cell(fila, 1, titulo).font = Font(
        bold=True, size=12, color=color, name='Arial')
    fila += 1

ws1.column_dimensions['A'].width = 80
ws1.column_dimensions['B'].width = 20

# ── HOJA 2: Clasificación ─────────────────────────────────────
ws2 = wb.create_sheet('Clasificación')
headers_clf = ['Modelo', 'Tipo', 'Accuracy', 'F1-macro',
               'F1-weighted', 'ROC-AUC',
               'CV F1-macro', 'CV Std', 'Gap', 'Recomendado']
for j, h in enumerate(headers_clf, 1):
    c = ws2.cell(1, j, h)
    estilo_header(c)

for i, nombre in enumerate(res_clf, 2):
    r     = res_clf[nombre]
    marca = 'SÍ ★' if nombre == mejor_clf else ''
    fill  = PatternFill('solid', start_color=VERDE) \
            if nombre == mejor_clf else None
    fila_data = [
        nombre, 'Ensemble',
        round(r['accuracy'], 4),
        round(r['f1_macro'], 4),
        round(r['f1_weighted'], 4),
        round(r['roc_auc'], 4),
        round(r['cv_mean'], 4),
        round(r['cv_std'], 4),
        round(r['gap'], 4),
        marca
    ]
    for j, val in enumerate(fila_data, 1):
        c = ws2.cell(i, j, val)
        estilo_celda(c, bold=(nombre == mejor_clf))
        if fill:
            c.fill = fill

ws2.cell(6, 1, 'DETALLE POR CLASE').font = Font(
    bold=True, size=11, color=AZUL_OSC, name='Arial')
headers_rep = ['Modelo', 'Clase', 'Precision',
               'Recall', 'F1-score', 'Support']
for j, h in enumerate(headers_rep, 1):
    c = ws2.cell(7, j, h)
    estilo_header(c, AZUL_MED)

fila_rep = 8
for nombre in res_clf:
    rep = reports_clf[nombre]
    for clase in le.classes_:
        if clase in rep:
            m = rep[clase]
            fila_data = [
                nombre, clase,
                round(m['precision'], 4),
                round(m['recall'], 4),
                round(m['f1-score'], 4),
                int(m['support'])
            ]
            color_clase = (ROJO_CLA if clase == 'ACEPTABLE'
                           else AZUL_CLA)
            for j, val in enumerate(fila_data, 1):
                c = ws2.cell(fila_rep, j, val)
                estilo_celda(c)
                c.fill = PatternFill('solid',
                                     start_color=color_clase)
            fila_rep += 1

autofit(ws2)

# ── HOJA 3: Regresión ────────────────────────────────────────
ws3 = wb.create_sheet('Regresión ICA')
headers_reg = ['Modelo', 'RMSE', 'MAE', 'R²',
               'MAPE (%)', 'RMSE Zona Crítica', 'Recomendado']
for j, h in enumerate(headers_reg, 1):
    c = ws3.cell(1, j, h)
    estilo_header(c)

for i, nombre in enumerate(res_reg, 2):
    r     = res_reg[nombre]
    marca = 'SÍ ★' if nombre == mejor_reg else ''
    fill  = PatternFill('solid', start_color=VERDE) \
            if nombre == mejor_reg else None
    rmse_c = round(r['rmse_crit'], 4) \
             if not np.isnan(r['rmse_crit']) else 'N/A'
    fila_data = [
        nombre,
        round(r['rmse'], 4),
        round(r['mae'],  4),
        round(r['r2'],   4),
        round(r['mape'], 2),
        rmse_c, marca
    ]
    for j, val in enumerate(fila_data, 1):
        c = ws3.cell(i, j, val)
        estilo_celda(c, bold=(nombre == mejor_reg))
        if fill:
            c.fill = fill

autofit(ws3)

# ── HOJA 4: Anomalías ─────────────────────────────────────────
ws4 = wb.create_sheet('Detección Anomalías')
ws4['A1'] = (f'Features usadas: {len(FEATURES_ANOM)} '
             f'(turbidez excluida — sin fuga de etiqueta)')
ws4['A1'].font = Font(size=10, italic=True,
                      color='C0392B', name='Arial')

headers_an = ['Modelo', 'Tipo', 'Avg Precision',
              'ROC-AUC', 'F1-Anomalía',
              'VP', 'FP', 'FN', 'VN', 'Recomendado']
for j, h in enumerate(headers_an, 1):
    c = ws4.cell(2, j, h)
    estilo_header(c, '6B4C8A')

for i, nombre in enumerate(res_anom, 3):
    r     = res_anom[nombre]
    marca = 'SÍ ★' if nombre == mejor_anom else ''
    fill  = PatternFill('solid', start_color=VERDE) \
            if nombre == mejor_anom else None
    cm    = r['cm']
    tn, fp, fn, tp = cm.ravel() if cm.size == 4 else (0,0,0,0)
    fila_data = [
        nombre, r['tipo'],
        round(r['ap'],  4),
        round(r['roc'], 4),
        round(r['f1'],  4),
        int(tp), int(fp), int(fn), int(tn),
        marca
    ]
    for j, val in enumerate(fila_data, 1):
        c = ws4.cell(i, j, val)
        estilo_celda(c, bold=(nombre == mejor_anom))
        if fill:
            c.fill = fill

autofit(ws4)

# ── HOJA 5: Features ─────────────────────────────────────────
ws5 = wb.create_sheet('Importancia Features')
headers_fi = ['Ranking', 'Feature', 'Importancia',
              'Importancia (%)']
for j, h in enumerate(headers_fi, 1):
    c = ws5.cell(1, j, h)
    estilo_header(c)

feat_imp_all = pd.Series(
    rf_model.feature_importances_, index=FEATURES
).sort_values(ascending=False)

for i, (feat, imp) in enumerate(feat_imp_all.items(), 2):
    fila_data = [i-1, feat,
                 round(imp, 6),
                 f"{imp*100:.3f}%"]
    color_f = AZUL_CLA if i <= 12 else 'FFFFFF'
    for j, val in enumerate(fila_data, 1):
        c = ws5.cell(i, j, val)
        estilo_celda(c, center=(j != 2))
        c.fill = PatternFill('solid', start_color=color_f)

autofit(ws5)

# ── HOJA 6: Eventos anómalos verificados ─────────────────────
ws6 = wb.create_sheet('Eventos Anómalos')
scores_all_full = -iso.score_samples(
    scaler_anom.transform(df_fe[FEATURES_ANOM].values))
scores_all_full = ((scores_all_full - scores_all_full.min()) /
                   (scores_all_full.max() -
                    scores_all_full.min() + 1e-9))
df_fe['score_iso'] = scores_all_full

headers_ev = ['Fecha_Hora', 'Turbidez_NTU',
              'Score_Anomalía', 'Clasificacion_Real',
              'Nivel_Alerta']
for j, h in enumerate(headers_ev, 1):
    c = ws6.cell(1, j, h)
    estilo_header(c, '6B4C8A')

eventos = df_fe[df_fe['anomalia'] == 1].copy()
eventos['nivel'] = eventos['score_iso'].apply(
    lambda s: 'ROJO' if s >= 0.65 else 'AMARILLO')

for i, (_, row) in enumerate(eventos.iterrows(), 2):
    color_ev = ('FADBD8' if row['nivel'] == 'ROJO'
                else 'FEF9E7')
    fila_data = [
        str(row['Fecha_Hora'])[:16],
        round(row['Turbidez_NTU'], 2),
        round(row['score_iso'], 4),
        row['Clasificacion'],
        row['nivel']
    ]
    for j, val in enumerate(fila_data, 1):
        c = ws6.cell(i, j, val)
        estilo_celda(c)
        c.fill = PatternFill('solid', start_color=color_ev)

autofit(ws6)

RUTA_EXCEL = 'reports/metrics/resultados_modelos.xlsx'
wb.save(RUTA_EXCEL)
print(f"  ✓ Excel guardado: {RUTA_EXCEL}")

print(f"\n{'='*65}")
print("  RESUMEN FINAL — ARCHIVOS EXPORTADOS")
print(f"{'='*65}")
print(f"\n  EXCEL:")
print(f"  reports/metrics/resultados_modelos.xlsx")
print(f"    · Hoja 1: Resumen ejecutivo")
print(f"    · Hoja 2: Clasificación (tabla + detalle por clase)")
print(f"    · Hoja 3: Regresión ICA")
print(f"    · Hoja 4: Detección de anomalías (sin turbidez)")
print(f"    · Hoja 5: Importancia de features")
print(f"    · Hoja 6: Eventos anómalos verificados")
print(f"\n  FIGURAS PNG (reports/figures/):")
figuras = [
    "fig1_comparacion_clasificacion.png",
    "fig2_matrices_confusion.png",
    "fig3_importancia_features.png",
    "fig4_prediccion_vs_real.png",
    "fig5_scatter_prediccion.png",
    "fig6_anomalias_tiempo.png",
    "fig7_precision_recall.png",
    "fig8_confusion_anomalias.png",
]
for f in figuras:
    print(f"  · {f}")
print(f"\n  MODELOS (data/models/):")
modelos_guardados = [
    'rf_clasificador.pkl', 'xgb_regresor.pkl',
    'isolation_forest.pkl', 'rf_anomalias.pkl',
    'scaler_clf.pkl', 'scaler_reg.pkl',
    'scaler_anom.pkl', 'label_encoder.pkl'
]
for m in modelos_guardados:
    print(f"  · {m}")
print(f"\n  ✓ PIPELINE COMPLETO")
print(f"  Ejecuta ahora: python src/alerts.py")